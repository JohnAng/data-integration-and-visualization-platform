"""
ETL transformer — read three heterogeneous bibliographic sources,
normalise them, resolve entities and emit the nine Polars frames
that the exporter writes to disk.

Phase I of the pipeline:

1. *Extract* — read DBLP article and inproceedings CSVs, the iCore
   conference ranking, the Kaggle Scimago journal ranking, and the
   Field-of-Research category lookup.
2. *Transform* — normalise titles / acronyms, expand abbreviations
   (``Trans. -> Transactions`` …), split pipe-separated author lists,
   validate against ``year ∈ [1900, 2030]`` and a non-null title.
3. *Resolve* — case-insensitive exact join of DBLP ``booktitle``
   against iCore ``Acronym`` (conferences) and a
   ``rapidfuzz.token_set_ratio >= 85`` over normalised titles
   (journals).
4. *Assign surrogate keys* — integer PKs for every lookup row.
5. *Quarantine* — every dropped row is written to the rejection log
   with reason and raw payload, never silently lost.

The output is a ``dict[str, pl.DataFrame]`` keyed by table name,
ready to be serialised by ``etl.exporter.DataExporter``.
"""

import json
import os
import re
import time
from pathlib import Path

import openpyxl
import polars as pl
from rapidfuzz import fuzz, process

REJECTION_COLUMNS: list[str] = ["source_file", "source_id", "reason", "raw_row"]


JOURNAL_TOKEN_EXPANSIONS: dict[str, str] = {
    "trans": "transactions",
    "j": "journal",
    "jrnl": "journal",
    "inf": "information",
    "comp": "computer",
    "comput": "computing",
    "knowl": "knowledge",
    "eng": "engineering",
    "sci": "science",
    "tech": "technology",
    "syst": "systems",
    "conf": "conference",
    "proc": "proceedings",
    "int": "international",
    "intl": "international",
    "symp": "symposium",
    "comm": "communications",
    "commun": "communications",
    "sec": "security",
    "lett": "letters",
    "math": "mathematics",
    "ann": "annals",
    "rev": "review",
    "res": "research",
    "appl": "applied",
    "theor": "theoretical",
    "softw": "software",
    "netw": "networks",
    "stat": "statistics",
    "biol": "biological",
    "med": "medicine",
    "amer": "american",
    "phys": "physics",
    "chem": "chemistry",
    "ind": "industrial",
    "manuf": "manufacturing",
    "auto": "automation",
    "autom": "automation",
    "mgmt": "management",
    "educ": "education",
    "psychol": "psychology",
    "soc": "society",
    "comb": "combinatoria",
    "cybern": "cybernetica",
    "emb": "embedded",
    "fund": "fundamenta",
    "informaticae": "informaticae",
    "informatik": "informatik",
    "numer": "numerique",
    "ling": "linguistic",
    "lit": "literary",
    "automat": "automation",
    "control": "control",
    "instrum": "instrumentation",
    "aero": "aerospace",
    "neur": "neural",
    "artif": "artificial",
    "intell": "intelligence",
    "alg": "algorithms",
    "discr": "discrete",
    "geom": "geometry",
    "logic": "logic",
    "anal": "analysis",
    "graph": "graphics",
    "vis": "visualization",
    "simul": "simulation",
    "model": "modelling",
    "dist": "distributed",
    "min": "mining",
    "decis": "decision",
    "exp": "experimental",
    "pract": "practice",
    "exper": "experience",
    "lang": "languages",
    "transl": "translation",
}


ACRONYM_STOP_WORDS: frozenset[str] = frozenset({
    "a", "an", "and", "or", "in", "on", "of", "the", "for", "to", "from",
    "with", "by", "at", "into", "&",
})

PUBLISHER_PREFIXES: tuple[str, ...] = (
    "acm/ieee", "ieee/acm", "ieee", "acm", "springer", "elsevier", "wiley",
    "sage", "oxford", "cambridge", "kluwer", "siam",
)


def _normalize_journal_title(title: str) -> str:
    """
    Normalize a journal title for fuzzy comparison.

    Replaces the ampersand with the literal word 'and', lowercases,
    strips non-alphanumeric characters, splits into tokens, drops common
    salutation words, and expands BibTeX-style abbreviations.
    """
    lowered = title.lower().replace("&", " and ")
    stripped = re.sub(r"[^a-z0-9\s]", " ", lowered)
    tokens = [tok for tok in stripped.split() if tok]
    expanded = [JOURNAL_TOKEN_EXPANSIONS.get(tok, tok) for tok in tokens]
    return " ".join(expanded)


def _generate_journal_acronyms(title: str) -> list[str]:
    """
    Generate the set of plausible DBLP-style acronyms for a Kaggle journal
    title.

    Covers the most common naming conventions:
      * 'ACM Transactions on X Y Z' -> 'TXYZ', 'TOXYZ'
      * 'IEEE Transactions on X Y' -> 'TXY', 'TOXY'
      * 'International Journal of X Y' -> 'IJXY'
      * 'Journal of X Y' -> 'JXY'
      * 'Proceedings of the X Y' -> 'PXY'
    Returns a deduplicated list of uppercase acronyms with three or more
    characters.
    """
    if not title:
        return []

    cleaned = title.strip()
    lowered = cleaned.lower()
    for prefix in PUBLISHER_PREFIXES:
        if lowered.startswith(prefix + " "):
            cleaned = cleaned[len(prefix) + 1:]
            lowered = cleaned.lower()
            break

    candidates: set[str] = set()

    transactions_match = re.match(
        r"^(?P<head>transactions?|trans\.?)\s+(?:on\s+|of\s+)?(?P<rest>.+)$",
        cleaned, re.IGNORECASE,
    )
    journal_match = re.match(
        r"^(?:(?P<inter>international)\s+)?journal\s+(?:on\s+|of\s+|for\s+)?(?P<rest>.+)$",
        cleaned, re.IGNORECASE,
    )
    proceedings_match = re.match(
        r"^proceedings?\s+(?:of\s+(?:the\s+)?|on\s+)?(?P<rest>.+)$",
        cleaned, re.IGNORECASE,
    )

    def _initials(rest: str) -> str:
        words = re.findall(r"\b[A-Za-z]+\b", rest)
        significant = [w for w in words if w.lower() not in ACRONYM_STOP_WORDS]
        return "".join(w[0].upper() for w in significant)

    if transactions_match:
        rest = transactions_match.group("rest")
        ini = _initials(rest)
        if ini:
            candidates.add("T" + ini)
            candidates.add("TO" + ini)
    if journal_match:
        rest = journal_match.group("rest")
        ini = _initials(rest)
        if ini:
            if journal_match.group("inter"):
                candidates.add("IJ" + ini)
                candidates.add("IJO" + ini)
            else:
                candidates.add("J" + ini)
                candidates.add("JO" + ini)
    if proceedings_match:
        rest = proceedings_match.group("rest")
        ini = _initials(rest)
        if ini:
            candidates.add("P" + ini)
            candidates.add("PO" + ini)

    return [c for c in candidates if len(c) >= 3]


def _looks_like_acronym(text: str) -> bool:
    """
    Heuristic check for whether a DBLP journal name is an acronym rather
    than a full title.
    """
    if not text:
        return False
    stripped = text.strip()
    if " " in stripped:
        return False
    return 2 <= len(stripped) <= 10 and stripped.replace(".", "").isalnum()


def _canonicalize_booktitle(booktitle: str) -> str:
    """
    Map a raw DBLP booktitle to the canonical parent conference title.

    The same canonical conference shows up in DBLP under many surface
    forms: workshops ('ADMS@VLDB'), session splits ('HICSS (1)'), volumes
    ('ICPP, Vol. 3'), workshop tracks ('EDBT/ICDT Workshops'), poster
    sessions, and so on. This function collapses all of them to the
    parent conference title so that downstream loaders deduplicate
    correctly, articles inherit their parent conference's iCore rank,
    and analytics aggregate at a meaningful granularity.

    Returns an empty string for empty input.
    """
    if booktitle is None:
        return ""
    text = booktitle.strip()
    if not text:
        return ""

    if "@" in text:
        text = text.split("@", 1)[1].strip()

    text = re.sub(r"\s*\([^)]*\)\s*", " ", text)
    text = re.sub(r"\s*,\s*(vol\.?|volume|part|no\.?|issue)\s+[a-z0-9\-]+\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(ph\.?\s*d\.?\s+)?workshops?\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+poster\s+sessions?\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+demonstrations?\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(industrial|industry)\s+track\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+tutorials?\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+companion\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+adjunct\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+best\s+of\s+.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+extended\s+abstracts?\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(spring|fall|summer|winter|autumn)\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(asia|europe|americas|pacific|atlantic)\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(short|long|short\s+papers?|long\s+papers?|research\s+papers?)\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(symposium\s+poster\s+session)\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(satellite\s+events?)\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(doctoral\s+(consortium|symposium))\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(supplement|supplemental\s+proceedings?)\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+\d{4}\s*$", "", text)
    text = re.sub(r"\s+(in\s+conjunction\s+with).*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(co-located\s+with).*$", "", text, flags=re.IGNORECASE)

    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_conf_acronym(booktitle: str) -> str:
    """
    Normalize a canonicalized booktitle to an uppercase acronym key for
    direct lookup against iCore.
    """
    canonical = _canonicalize_booktitle(booktitle)
    return canonical.upper() if canonical else ""


class DataTransformer:
    """
    Executes the Extract and Transform phases of the ETL pipeline.

    Reads the three raw bibliographic sources (DBLP articles, DBLP
    inproceedings, Kaggle journal rankings, iCore26 conference rankings)
    and produces the in-memory DataFrames that will be exported to CSV
    by the DataExporter and then loaded into the MySQL star schema.

    Entity resolution between DBLP and the ranking sources follows a
    hybrid strategy:
      * Conferences: DBLP booktitle is treated as an acronym and joined
        exactly against the iCore Acronym column after normalization.
      * Journals: DBLP abbreviated titles are matched against Kaggle full
        titles via rapidfuzz token_set_ratio with a configurable cutoff.

    The optional sample_rows parameter limits how many rows are read from
    each DBLP source file, enabling fast iterative development on a small
    subset of the full ~11M-row dataset.
    """

    def __init__(
        self,
        sample_rows: int | None = None,
        journal_fuzzy_cutoff: int = 80,
        conference_fuzzy_cutoff: int = 88,
        data_root: Path | None = None,
    ) -> None:
        """
        Initialize the transformer with source paths and tuning knobs.

        Parameters
        ----------
        sample_rows:
            If provided, only the first N rows of each DBLP CSV are read.
        journal_fuzzy_cutoff:
            Minimum rapidfuzz token_set_ratio score (0-100) required to
            accept a Kaggle journal as a match for a DBLP journal title.
        conference_fuzzy_cutoff:
            Minimum token_set_ratio for the fuzzy fallback that attempts
            to match DBLP booktitles against iCore full titles when the
            direct acronym lookup fails. Higher than the journal cutoff
            because conference titles are typically less abbreviated.
        data_root:
            Optional override for the directory that contains the three
            source subfolders (dblp_dataset, journal_ranking_data_raw,
            icore26_data). Tests use this to point at a fixture tree.
        """
        if data_root is None:
            data_root = Path(__file__).resolve().parent.parent.parent.parent / "data"
        self.base_dir = data_root
        self.dblp_dir = self.base_dir / "dblp_dataset"
        self.kaggle_dir = self.base_dir / "journal_ranking_data_raw"
        self.icore_dir = self.base_dir / "icore26_data"
        self.sample_rows = sample_rows
        self.journal_fuzzy_cutoff = journal_fuzzy_cutoff
        self.conference_fuzzy_cutoff = conference_fuzzy_cutoff
        self.match_stats: dict[str, int] = {}
        self._rejections: list[dict] = []
        self._booktitle_canonical: dict[str, str] = {}

    def _record_rejections(
        self,
        invalid_rows: pl.DataFrame,
        source_file: str,
        reason: str,
        id_column: str = "id",
    ) -> None:
        """
        Append each row of invalid_rows to the rejection log with the given
        reason. The raw row is serialized to a JSON string for forensic
        inspection inside the Rejection_Log table.
        """
        if invalid_rows.height == 0:
            return
        for row in invalid_rows.to_dicts():
            raw_source_id = row.get(id_column)
            self._rejections.append({
                "source_file": source_file,
                "source_id": str(raw_source_id) if raw_source_id is not None else None,
                "reason": reason,
                "raw_row": json.dumps(row, ensure_ascii=False, default=str),
            })

    def _clean_and_validate_articles(
        self,
        articles: pl.DataFrame,
        source_file: str,
        venue_column: str,
    ) -> pl.DataFrame:
        """
        Apply article-level cleanup and validation.

        Salvageable defects are silently corrected so the row can still be
        loaded: leading/trailing whitespace is stripped, internal multiple
        spaces are collapsed, oversized text fields are truncated to the
        column limits, and out-of-range years are null'd out. Only rows
        with no recoverable identity (missing venue or missing title) are
        routed to the rejection log.
        """
        articles = articles.with_columns([
            pl.col(venue_column)
                .cast(pl.Utf8, strict=False)
                .str.strip_chars()
                .str.replace_all(r"\s+", " "),
            pl.col("title")
                .cast(pl.Utf8, strict=False)
                .str.strip_chars()
                .str.replace_all(r"\s+", " "),
            pl.col("pages").cast(pl.Utf8, strict=False).str.strip_chars(),
            pl.col("url").cast(pl.Utf8, strict=False).str.strip_chars(),
            pl.col("id").cast(pl.Utf8, strict=False).str.strip_chars(),
        ])

        missing_venue = articles[venue_column].is_null() | (articles[venue_column] == "")
        self._record_rejections(articles.filter(missing_venue), source_file, f"missing_{venue_column}")
        articles = articles.filter(~missing_venue)

        missing_title = articles["title"].is_null() | (articles["title"] == "")
        self._record_rejections(articles.filter(missing_title), source_file, "missing_title")
        articles = articles.filter(~missing_title)

        year_as_integer = articles["year"].cast(pl.Int32, strict=False)
        articles = articles.with_columns([
            pl.col("title").str.slice(0, 500),
            pl.col("pages").str.slice(0, 100),
            pl.col("url").str.slice(0, 500),
            pl.when((year_as_integer >= 1900) & (year_as_integer <= 2100))
                .then(year_as_integer)
                .otherwise(None)
                .alias("year"),
        ])

        return articles

    def extract_for_categories(self) -> pl.DataFrame:
        """
        Load icoreCategories.xlsx and return one row per Field of Research
        category code, with description and self-referential parent_code.

        The xlsx is structured with one column of codes and one column of
        descriptions. Codes are either four-digit parents (4601), six-digit
        children (460101 maps to parent 4601), or the special string 'CSE'
        for Computer Systems Engineering. Whitespace inside descriptions
        is normalized to a single space.
        """
        xlsx_path = self.icore_dir / "icoreCategories.xlsx"
        if not xlsx_path.exists():
            raise FileNotFoundError(f"Missing source file: {xlsx_path}")

        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]

        rows: list[dict] = []
        seen: set[str] = set()
        for row in worksheet.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            code = str(row[0]).strip()
            raw_desc = row[1] if len(row) > 1 else None
            description = re.sub(r"\s+", " ", str(raw_desc)).strip() if raw_desc is not None else ""
            if not code or not description or code in seen:
                continue
            parent: str | None = None
            if code.isdigit() and len(code) == 6:
                parent = code[:4]
            rows.append({"code": code, "description": description, "parent_code": parent})
            seen.add(code)
        workbook.close()

        return pl.DataFrame(rows, schema={
            "code": pl.Utf8,
            "description": pl.Utf8,
            "parent_code": pl.Utf8,
        })

    def extract_rejections(self) -> pl.DataFrame:
        """
        Materialize the accumulated rejection log as a Polars DataFrame
        matching the Rejection_Log table schema (minus auto-generated
        columns).
        """
        if not self._rejections:
            return pl.DataFrame(
                schema={
                    "source_file": pl.Utf8,
                    "source_id": pl.Utf8,
                    "reason": pl.Utf8,
                    "raw_row": pl.Utf8,
                }
            )
        return pl.DataFrame(self._rejections).select(REJECTION_COLUMNS)

    def _read_dblp_source(self, filename: str, columns: list[str]) -> pl.DataFrame:
        """
        Read a DBLP semicolon-separated source file with strict parsing.

        Truncates ragged lines and skips parse errors to survive the
        well-known DBLP CSV quirks. Honors the sample_rows setting.
        """
        file_path = self.dblp_dir / filename
        if not file_path.exists():
            raise FileNotFoundError(f"Missing source file: {file_path}")

        return pl.read_csv(
            file_path,
            separator=";",
            ignore_errors=True,
            truncate_ragged_lines=True,
            quote_char='"',
            columns=columns,
            n_rows=self.sample_rows,
        )

    def extract_authors(self) -> pl.DataFrame:
        """
        Extract, split, normalize, and deduplicate authors from both DBLP
        sources.

        Authors are stored in DBLP as '|'-delimited strings; they are
        exploded into one row each, trimmed, internal whitespace is
        collapsed, oversized names are truncated to the schema limit
        (255 characters), and exact duplicates are removed before assigning
        surrogate author_id values.
        """
        inproceedings_authors = self._read_dblp_source("input_inproceedings.csv", ["author"])
        article_authors = self._read_dblp_source("input_article.csv", ["author"])

        combined_author_rows = pl.concat([
            inproceedings_authors.select(pl.col("author")),
            article_authors.select(pl.col("author")),
        ])

        return (
            combined_author_rows
            .with_columns(pl.col("author").str.split("|"))
            .explode("author")
            .drop_nulls()
            .with_columns(
                pl.col("author")
                    .str.strip_chars()
                    .str.replace_all(r"\s+", " ")
            )
            .filter(pl.col("author") != "")
            .with_columns(pl.col("author").str.slice(0, 255))
            .with_columns(pl.col("author")
                    .str.normalize("NFKD")
                    .str.replace_all(r"\p{M}", "")
                    .str.to_lowercase()
                    .alias("author_dedup_key"))
            .unique(subset=["author_dedup_key"], maintain_order=True)
            .drop("author_dedup_key")
            .sort("author")
            .rename({"author": "author_name"})
            .with_row_index(name="author_id", offset=1)
            .select(["author_id", "author_name"])
        )

    def extract_journals(self) -> pl.DataFrame:
        """
        Extract distinct journals from DBLP and enrich them via fuzzy match
        against the Kaggle ranking source.

        Returns one row per distinct DBLP journal title, with all Kaggle
        metric columns populated when a match above the cutoff exists and
        NULL otherwise. Records match statistics on self.match_stats.
        """
        dblp_journal_column = self._read_dblp_source("input_article.csv", ["journal"])
        dblp_titles: list[str] = (
            dblp_journal_column
            .drop_nulls()
            .filter(pl.col("journal").str.strip_chars() != "")
            .with_columns(pl.col("journal")
                .str.normalize("NFKD")
                .str.replace_all(r"\p{M}", "")
                .str.to_lowercase()
                .alias("journal_dedup_key"))
            .unique(subset=["journal_dedup_key"], maintain_order=True)
            .drop("journal_dedup_key")
            .get_column("journal")
            .to_list()
        )

        kaggle_journals = pl.read_csv(
            self.kaggle_dir / "journal_ranking_data_raw.csv",
            separator=",",
            ignore_errors=True,
        ).select(
            pl.col("Title").alias("kaggle_title"),
            pl.col("Publisher").alias("publisher"),
            pl.col("Country").alias("country"),
            pl.col("Best Quartile").alias("best_quartile"),
            pl.col("Best Subject Area").alias("best_subject_area"),
            pl.col("SJR-index").cast(pl.Float64, strict=False).alias("sjr_index"),
            pl.col("CiteScore").cast(pl.Float64, strict=False).alias("citation_score"),
            pl.col("H-index").cast(pl.Int32, strict=False).alias("h_index"),
            pl.col("Total Docs.").cast(pl.Int32, strict=False).alias("total_documents"),
            pl.col("Total Docs. 3y").cast(pl.Int32, strict=False).alias("total_documents_3y"),
            pl.col("Total Refs.").cast(pl.Int32, strict=False).alias("total_references"),
            pl.col("Total Cites 3y").cast(pl.Int32, strict=False).alias("total_citations_3y"),
            pl.col("Citable Docs. 3y").cast(pl.Int32, strict=False).alias("citable_documents_3y"),
            pl.col("Cites/Doc. 2y").cast(pl.Float64, strict=False).alias("citations_per_document_2y"),
            pl.col("Refs./Doc.").cast(pl.Float64, strict=False).alias("references_per_document"),
        ).unique(subset=["kaggle_title"])

        kaggle_rows = kaggle_journals.to_dicts()
        kaggle_index: dict[str, dict] = {}
        kaggle_acronym_index: dict[str, dict] = {}
        for row in kaggle_rows:
            kaggle_title = row["kaggle_title"]
            if kaggle_title is None:
                continue
            key = _normalize_journal_title(kaggle_title)
            if key and key not in kaggle_index:
                kaggle_index[key] = row
            for acronym in _generate_journal_acronyms(kaggle_title):
                kaggle_acronym_index.setdefault(acronym, row)
        kaggle_keys = list(kaggle_index.keys())

        empty_metrics: dict = {
            field: None for field in (
                "publisher", "country", "best_quartile", "best_subject_area",
                "sjr_index", "citation_score", "h_index", "total_documents",
                "total_documents_3y", "total_references", "total_citations_3y",
                "citable_documents_3y", "citations_per_document_2y", "references_per_document",
            )
        }

        matched_rows: list[dict] = []
        exact_hits = 0
        acronym_hits = 0
        fuzzy_hits = 0
        misses = 0

        for dblp_title in dblp_titles:
            normalized = _normalize_journal_title(dblp_title)
            metrics: dict | None = None

            if normalized and normalized in kaggle_index:
                metrics = kaggle_index[normalized]
                exact_hits += 1
            elif _looks_like_acronym(dblp_title):
                acronym_key = dblp_title.strip().replace(".", "").upper()
                if acronym_key in kaggle_acronym_index:
                    metrics = kaggle_acronym_index[acronym_key]
                    acronym_hits += 1
                else:
                    misses += 1
            elif normalized:
                hit = process.extractOne(
                    normalized,
                    kaggle_keys,
                    scorer=fuzz.token_set_ratio,
                    score_cutoff=self.journal_fuzzy_cutoff,
                )
                if hit is not None:
                    matched_key = hit[0]
                    metrics = kaggle_index[matched_key]
                    fuzzy_hits += 1
                else:
                    misses += 1
            else:
                misses += 1

            row_metrics = metrics if metrics is not None else empty_metrics
            matched_rows.append({
                "title": dblp_title,
                "publisher": row_metrics["publisher"],
                "country": row_metrics["country"],
                "best_quartile": row_metrics["best_quartile"],
                "best_subject_area": row_metrics["best_subject_area"],
                "sjr_index": row_metrics["sjr_index"],
                "citation_score": row_metrics["citation_score"],
                "h_index": row_metrics["h_index"],
                "total_documents": row_metrics["total_documents"],
                "total_documents_3y": row_metrics["total_documents_3y"],
                "total_references": row_metrics["total_references"],
                "total_citations_3y": row_metrics["total_citations_3y"],
                "citable_documents_3y": row_metrics["citable_documents_3y"],
                "citations_per_document_2y": row_metrics["citations_per_document_2y"],
                "references_per_document": row_metrics["references_per_document"],
            })

        self.match_stats["journals_total"] = len(dblp_titles)
        self.match_stats["journals_exact"] = exact_hits
        self.match_stats["journals_acronym"] = acronym_hits
        self.match_stats["journals_fuzzy"] = fuzzy_hits
        self.match_stats["journals_unmatched"] = misses

        return (
            pl.DataFrame(matched_rows)
            .with_columns([
                pl.col("title").str.strip_chars().str.replace_all(r"\s+", " ").str.slice(0, 500),
                pl.col("publisher").str.strip_chars().str.slice(0, 255),
                pl.col("country").str.strip_chars().str.slice(0, 100),
                pl.col("best_quartile").str.strip_chars().str.slice(0, 2),
                pl.col("best_subject_area").str.strip_chars().str.slice(0, 100),
            ])
            .with_row_index(name="journal_id", offset=1)
            .select([
                "journal_id", "title", "publisher", "country", "best_quartile",
                "best_subject_area", "sjr_index", "citation_score", "h_index",
                "total_documents", "total_documents_3y", "total_references", "total_citations_3y",
                "citable_documents_3y", "citations_per_document_2y", "references_per_document",
            ])
        )

    def extract_conferences(self) -> pl.DataFrame:
        """
        Extract distinct CANONICAL conferences from DBLP and enrich them
        via acronym + fuzzy match against the iCore26 ranking source.

        The raw DBLP booktitle namespace contains many surface variants
        of the same parent conference (workshops, session splits, vol/part
        suffixes, etc.). Those are collapsed via _canonicalize_booktitle
        so that one Lookup_Conference row represents the parent
        conference and articles from any of its variants inherit the
        same iCore rank and primary_for. The original raw booktitle to
        canonical title mapping is retained on self._booktitle_canonical
        for the fact loader.
        """
        dblp_booktitle_column = self._read_dblp_source("input_inproceedings.csv", ["booktitle"])
        raw_titles: list[str] = (
            dblp_booktitle_column
            .drop_nulls()
            .filter(pl.col("booktitle").str.strip_chars() != "")
            .unique(subset=["booktitle"])
            .get_column("booktitle")
            .to_list()
        )

        self._booktitle_canonical: dict[str, str] = {}
        canonical_unique: list[str] = []
        canonical_keys_seen: set[str] = set()
        for raw in raw_titles:
            canonical = _canonicalize_booktitle(raw)
            if not canonical:
                continue
            self._booktitle_canonical[raw] = canonical
            canonical_key = canonical.lower()
            if canonical_key not in canonical_keys_seen:
                canonical_keys_seen.add(canonical_key)
                canonical_unique.append(canonical)

        icore_conferences = pl.read_csv(
            self.icore_dir / "iCore26_KilledColumnsForLoading.csv",
            separator=",",
            ignore_errors=True,
        ).select(
            pl.col(" Title").alias("icore_title"),
            pl.col("Acronym").alias("acronym"),
            pl.col("Rank").alias("rank_value"),
            pl.col("PrimaryFoR").cast(pl.Utf8, strict=False).alias("primary_for"),
        )

        icore_rows = icore_conferences.to_dicts()
        icore_by_acronym: dict[str, dict] = {}
        icore_by_title: dict[str, dict] = {}
        for row in icore_rows:
            acronym = row["acronym"]
            if acronym is not None:
                acro_key = acronym.strip().upper()
                if acro_key and acro_key not in icore_by_acronym:
                    icore_by_acronym[acro_key] = row
            title = row["icore_title"]
            if title is not None:
                title_key = _normalize_journal_title(title)
                if title_key and title_key not in icore_by_title:
                    icore_by_title[title_key] = row
        icore_title_keys = list(icore_by_title.keys())

        matched_rows: list[dict] = []
        acronym_hits = 0
        fuzzy_hits = 0
        misses = 0

        for canonical in canonical_unique:
            metrics: dict | None = None

            acro_key = canonical.upper()
            if acro_key in icore_by_acronym:
                metrics = icore_by_acronym[acro_key]
                acronym_hits += 1
            else:
                title_key = _normalize_journal_title(canonical)
                if title_key:
                    hit = process.extractOne(
                        title_key,
                        icore_title_keys,
                        scorer=fuzz.token_set_ratio,
                        score_cutoff=self.conference_fuzzy_cutoff,
                    )
                    if hit is not None:
                        metrics = icore_by_title[hit[0]]
                        fuzzy_hits += 1
                    else:
                        misses += 1
                else:
                    misses += 1

            matched_rows.append({
                "title": canonical,
                "acronym": metrics["acronym"] if metrics else None,
                "rank_value": metrics["rank_value"] if metrics else None,
                "primary_for": metrics["primary_for"] if metrics else None,
            })

        self.match_stats["conferences_raw_dblp"] = len(raw_titles)
        self.match_stats["conferences_canonical"] = len(canonical_unique)
        self.match_stats["conferences_acronym"] = acronym_hits
        self.match_stats["conferences_fuzzy"] = fuzzy_hits
        self.match_stats["conferences_unmatched"] = misses

        return (
            pl.DataFrame(matched_rows)
            .with_columns([
                pl.col("title").str.strip_chars().str.replace_all(r"\s+", " ").str.slice(0, 500),
                pl.col("acronym").str.strip_chars().str.slice(0, 50),
                pl.col("rank_value").str.strip_chars().str.slice(0, 50),
                pl.col("primary_for").str.strip_chars().str.slice(0, 20),
            ])
            .with_row_index(name="conference_id", offset=1)
            .select(["conference_id", "title", "acronym", "rank_value", "primary_for"])
        )

    def extract_fact_journals(self, lookup_journal_frame: pl.DataFrame) -> pl.DataFrame:
        """
        Build the Fact_Journal_Article DataFrame by joining raw DBLP
         articles to the journal lookup on title. Invalid rows are routed
        to the rejection log; surviving rows are returned. source_id is
        retained for bridge resolution and data lineage.
        """
        columns_to_read = ["id", "title", "year", "pages", "url", "journal"]
        raw_articles = self._read_dblp_source("input_article.csv", columns_to_read)
        raw_articles = self._clean_and_validate_articles(raw_articles, "input_article.csv", venue_column="journal")

        return (
            raw_articles
            .with_columns(pl.col("journal")
                .str.normalize("NFKD")
                .str.replace_all(r"\p{M}", "")
                .str.to_lowercase()
                .alias("journal_match_key"))
            .join(
                lookup_journal_frame
                    .select(["journal_id", "title"])
                    .with_columns(pl.col("title")
                    .str.normalize("NFKD")
                    .str.replace_all(r"\p{M}", "")
                    .str.to_lowercase()
                    .alias("journal_match_key"))
                    .select(["journal_id", "journal_match_key"]),
                on="journal_match_key",
                how="inner",
            )
            .drop("journal_match_key")
            .with_row_index(name="article_id", offset=1)
            .select([
                "article_id",
                pl.col("id").cast(pl.Utf8).alias("source_id"),
                "title",
                pl.col("year").cast(pl.Int32, strict=False),
                "pages",
                "url",
                "journal_id",
            ])
        )

    def extract_fact_conferences(self, lookup_conference_frame: pl.DataFrame) -> pl.DataFrame:
        """
        Build the Fact_Conference_Article DataFrame by joining raw DBLP
        inproceedings rows to the conference lookup on the CANONICAL
        booktitle. Raw booktitle variants such as workshops and session
        splits are mapped to their parent conference via the
        _booktitle_canonical lookup populated by extract_conferences,
        so each article inherits its parent's rank and primary_for.
        """
        columns_to_read = ["id", "title", "year", "pages", "url", "booktitle"]
        raw_inproceedings = self._read_dblp_source("input_inproceedings.csv", columns_to_read)
        raw_inproceedings = self._clean_and_validate_articles(raw_inproceedings, "input_inproceedings.csv", venue_column="booktitle")

        canonical_map = self._booktitle_canonical
        raw_inproceedings = raw_inproceedings.with_columns(
            pl.col("booktitle")
                .map_elements(lambda raw_booktitle: canonical_map.get(raw_booktitle, ""), return_dtype=pl.Utf8)
                .alias("canonical_booktitle")
        ).filter(pl.col("canonical_booktitle") != "")

        return (
            raw_inproceedings
            .with_columns(pl.col("canonical_booktitle")
                .str.normalize("NFKD")
                .str.replace_all(r"\p{M}", "")
                .str.to_lowercase()
                .alias("conference_match_key"))
            .join(
                lookup_conference_frame
                    .select(["conference_id", "title"])
                    .with_columns(pl.col("title")
                    .str.normalize("NFKD")
                    .str.replace_all(r"\p{M}", "")
                    .str.to_lowercase()
                    .alias("conference_match_key"))
                    .select(["conference_id", "conference_match_key"]),
                on="conference_match_key",
                how="inner",
            )
            .drop("conference_match_key")
            .with_row_index(name="article_id", offset=1)
            .select([
                "article_id",
                pl.col("id").cast(pl.Utf8).alias("source_id"),
                "title",
                pl.col("year").cast(pl.Int32, strict=False),
                "pages",
                "url",
                "conference_id",
            ])
        )

    def extract_bridge_tables(
        self,
        filename: str,
        fact_frame: pl.DataFrame,
        lookup_author_frame: pl.DataFrame,
    ) -> pl.DataFrame:
        """
        Resolve the N:M relationship between articles and authors for one
        DBLP source file, returning (article_id, author_id) pairs.
        """
        source_rows = self._read_dblp_source(filename, ["id", "author"]).drop_nulls(subset=["author"])

        exploded_author_rows = (
            source_rows
            .with_columns(pl.col("author").str.split("|"))
            .explode("author")
            .with_columns(pl.col("author").str.strip_chars())
            .filter(pl.col("author") != "")
            .with_columns(pl.col("author")
                .str.normalize("NFKD")
                .str.replace_all(r"\p{M}", "")
                .str.to_lowercase()
                .alias("author_match_key"))
        )

        lookup_author_with_match_key = lookup_author_frame.with_columns(
            pl.col("author_name")
                .str.normalize("NFKD")
                .str.replace_all(r"\p{M}", "")
                .str.to_lowercase()
                .alias("author_match_key")
        ).select(["author_id", "author_match_key"])

        rows_with_author_ids = exploded_author_rows.join(
            lookup_author_with_match_key,
            on="author_match_key",
            how="inner",
        )

        return (
            rows_with_author_ids
            .with_columns(pl.col("id").cast(pl.Utf8))
            .join(fact_frame.select(["article_id", "source_id"]), left_on="id", right_on="source_id", how="inner")
            .select(["article_id", "author_id"])
            .unique()
        )

    def transform_all(self) -> dict[str, pl.DataFrame]:
        """
        Orchestrate the full extract+transform pipeline and return the
        eight DataFrames that the exporter will serialize to CSV. The
        Rejection_Log frame captures rows that failed validation across
        the pipeline.
        """
        lookup_for_category = self.extract_for_categories()
        lookup_authors = self.extract_authors()
        lookup_journals = self.extract_journals()
        lookup_conferences = self.extract_conferences()

        fact_journal = self.extract_fact_journals(lookup_journals)
        fact_conference = self.extract_fact_conferences(lookup_conferences)

        bridge_journal = self.extract_bridge_tables(
            "input_article.csv", fact_journal, lookup_authors,
        )
        bridge_conference = self.extract_bridge_tables(
            "input_inproceedings.csv", fact_conference, lookup_authors,
        )

        rejection_logs = self.extract_rejections()
        self.match_stats["rejections_total"] = rejection_logs.height

        return {
            "lookup_field_of_research_categories": lookup_for_category,
            "lookup_authors": lookup_authors,
            "lookup_journals": lookup_journals,
            "lookup_conferences": lookup_conferences,
            "fact_journal_articles": fact_journal,
            "fact_conference_articles": fact_conference,
            "bridge_journal_article_authors": bridge_journal,
            "bridge_conference_article_authors": bridge_conference,
            "rejection_logs": rejection_logs,
        }


if __name__ == "__main__":
    sample_env = os.environ.get("ETL_SAMPLE_ROWS")
    sample_rows = int(sample_env) if sample_env else None

    start = time.perf_counter()
    transformer = DataTransformer(sample_rows=sample_rows)
    datasets = transformer.transform_all()
    elapsed = time.perf_counter() - start

    label = f"sample={sample_rows}" if sample_rows else "full"
    print(f"\nETL transform completed in {elapsed:.2f}s ({label})\n")
    print("=== Match statistics ===")
    for name, value in transformer.match_stats.items():
        print(f"  {name:30s} {value}")
    print("\n=== Output table sizes ===")
    for table_name, dataframe in datasets.items():
        print(f"  {table_name:34s} rows={dataframe.height:>10d}  columns={dataframe.width:>3d}")
